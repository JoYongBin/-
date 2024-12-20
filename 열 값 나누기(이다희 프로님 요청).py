import pandas as pd                                          # 데이터 분석 사용
from openpyxl import load_workbook                           # 엑셀 파일을 수정하고 스타일 지정(사용하지 않을 경우, 기존 엑셀 스타일이 깨짐)
from openpyxl.utils.dataframe import dataframe_to_rows       # 회사ID가 32문자 이상일 경우 나눌 때 사용

# 엑셀 파일 읽기
file_path = 'C:/Downloads/Data_excel.xlsx'  # 원본 엑셀 파일 경로
df = pd.read_excel(file_path)

# '회사ID' 열을 문자열로 변환 및 공백 제거
df['회사ID'] = df['회사ID'].astype(str).str.replace(r'\s+', '', regex=True).str.strip()

# 회사 ID가 32글자씩 나누는 함수
def split_company_id(company_id):
    # 공백 제거 및 32글자씩 나누기
    company_id = company_id.replace('\n', '').replace('\r', '').strip()
    return [company_id[i:i+32] for i in range(0, len(company_id), 32)]

# 회사ID를 분리하여 리스트로 변환
df['회사ID'] = df['회사ID'].apply(split_company_id)

# 각 회사명에 대해 회사ID를 분리하여 새로운 행으로 확장
df_expanded = df.explode('회사ID')

# 빈 ID값이 있을 경우 해당 행 제거
df_expanded = df_expanded[df_expanded['회사ID'].str.strip() != '']

# 원본 엑셀 파일의 스타일 불러오기
wb = load_workbook(file_path)
ws = wb.active

# 처리된 데이터를 원본 엑셀 파일의 스타일을 유지하면서 새로운 파일에 저장
output_file_path = '파이프라인_fixed.xlsx'  # 결과 파일 경로

# 결과 엑셀 파일에 데이터 쓰기
for r_idx, row in enumerate(dataframe_to_rows(df_expanded, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# 결과 파일 저장
wb.save(output_file_path)

print(f" 기존 엑셀 원본 스타일을 유지한 데이터를 '{output_file_path}'에 엑셀 파일로 저장.")